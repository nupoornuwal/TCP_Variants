import subprocess
import os  
import sys
from operator import sub
import openpyxl
from openpyxl import Workbook
ns_command= "/course/cs4700f12/ns-allinone-2.35/bin/"
import pylab

variant= ["Reno","SACK"]
queue_variant=["DropTail","RED"]

for var in variant:
	for k in queue_variant :
		os.system("cd " +ns_command)
                os.system('ns '+'testqueue.tcl '+var+' '+k)	

def throughput(var,k):
        global thoughtput
        file='testqueue'+var+k+'.tr'
        #print 'FILE', file
        f=open(file)
        lines=f.readlines()
        #print lines
        f.close()
	start_received_time1=0.0
	start_received_time2=0.0
        received_packet1=0
        received_packet2=0
        start_time=0.0
        start_time1=0.0
        end_time1=0.0
        end_time2=0.0
        through_list=[]
        through_list1=[]
        x=0.25
        y=10.0
	time_list=[]
	time_list1=[]
        for k in lines:
                #print k
                n=k.split()
                if n[7]=="1":
			if n[0]=="+" and n[2]=="0" and n[4]=="tcp":
					if start_time==0.0:
						start_time=float(n[1])
		                                #print start_time		
                        if n[0]== "r" and  n[3]== "4" and n[4]=="tcp" :
                                        received_packet1=received_packet1+int(n[5])
                                        end_time1=float(n[1])
					#difference=end_time1-start_time
                                        #print 'end_time1',end_time1
                        if  float(n[1])>=x:
				
				difference=end_time1-start_time
				#print 'Difference',difference
				time_list.append(float(n[1]))
                                through=(received_packet1*8)/(difference*1000000)
                                through_list.append(through)
                                x+=0.25
				start_time=0.0
				received_packet1=0

                if n[7]=="2":
				
			if n[0]=="+" and n[2]=="1" and n[4]=="cbr":
                                        if start_time1==0.0:
                                                start_time1=float(n[1])
						#print start_time1
                        if n[0]== "r" and  n[3]== "5" and n[4]=="cbr" :
                                        received_packet2=received_packet2+int(n[5])
                                        end_time2=float(n[1])
                                        #print 'end_time2',end_time2
                        if float(n[1])>=y:
				difference1=end_time2-start_time1
				#print 'DIFFERENCE',difference1
				time_list1.append(float(n[1]))
                                through1=received_packet2*8/(difference1*1000000)
                                through_list1.append(through1)
                                y+=0.25
				start_time1=0.0
				received_packet2=0
				


        #print 'end_time1',end_time1
        #print 'received_time1',start_received_time1
        #print 'end_time2',end_time2
        #print 'received_time2',start_received_time2
        time1=(end_time1-start_received_time1)*1000000
        #print 'time',time1
        thoughput1=received_packet1*8/(float(time1))
        #print 'throughput',throughput1
        time2=(end_time2-start_received_time2)*1000000
        throughput2=received_packet2*8/(float(time2))
        #print 'throughlist',through_list, len(through_list)
	#print 'time1',time_list,len(time_list)
        #print 'throughlist1',through_list1
        return time_list,through_list,time_list1,through_list1

							   
for var in variant:
	for k in queue_variant:
		if var=="Reno" and k=="DropTail": 
			a=[]
			b=[]
			c=[]
			d=[]
			a,b,c,d=throughput(var,k)
			#pylab.figure(1)
			#pylab.plot(a,b,color='red')
			#pylab.plot(c,d,color='green')
			#pylab.grid()
			#pylab.show()
			wb = Workbook()
			ws = wb.create_sheet()
			#print a, len(a) 
			#print b,len(b)
			print c,len(c)			
	 		print d,len(d)
			for i in range(len(a)):
				ws['A'+str(i + 1)].value = a[i]
				ws['B'+str(i + 1)].value = b[i]
			for m in range(len(c)):
				ws['D'+str(m + 1)].value = c[m]
				ws['E'+str(m + 1)].value = d[m]
				wb.save( filename = 'Reno_Droptail.xlsx' )
		if var=="Reno" and k=="RED":
                        a=[]
                        b=[]
			c=[]
                        d=[]
                        a,b,c,d=throughput(var,k)
                        wb = Workbook()
                        ws = wb.create_sheet()
                        #print a, len(a)
                        #print b,len(b)
			#print c,len(c)
                        #print d,len(d)
			pylab.figure(1)
                        pylab.plot(a,b,color='blue')
                        pylab.plot(c,d,color='yellow')
                        pylab.grid()
			pylab.title('RED(Reno/SACK)')
			pylab.xlabel('Time(s)')
			pylab.ylabel('Throughput(Mbps)')
			#pylab.legend(['Reno_TCP(DropTail)','Reno_CBR(Droptail)','Reno_TCP(RED)','Reno_CBR(RED)'],2)
                        #pylab.show()

                        for i in range(len(a)):
                                ws['A'+str(i + 1)].value = a[i]
                                ws['B'+str(i + 1)].value = b[i]
				wb.save( filename = 'Reno_RED.xlsx' )
			for m in range(len(c)):
				ws['D'+str(m + 1)].value = c[m]
                                ws['E'+str(m + 1)].value = d[m]
                            	wb.save( filename = 'Reno_RED.xlsx' )
		if var=="SACK" and k=="DropTail":
                        a=[]
                        b=[]
                        c=[]
                        d=[]
                        a,b,c,d=throughput(var,k)
                        wb = Workbook()
                        ws = wb.create_sheet()
                        #print a, len(a)
                        #print b,len(b)
                        #print c,len(c)
                        #print d,len(d)
			#pylab.figure(1)
			#pylab.plot(a,b,color='blue')
                        #pylab.plot(c,d,color='cyan')
                        #pylab.grid()
                        #pylab.title('SACK(DropTail/RED)')
                        #pylab.xlabel('Time(s)')
                        #pylab.ylabel('Throughput(Mbps')
                        #pylab.legend(['Reno_TCP','Reno_CBR','SACK_TCP','SACK_CBR'],2)
                        #pylab.show()

                        for i in range(len(a)):
                                ws['A'+str(i + 1)].value = a[i]
                                ws['B'+str(i + 1)].value = b[i]
			for m in range(len(c)):
                                ws['D'+str(m + 1)].value = c[m]
                                ws['E'+str(m + 1)].value = d[m]
                                wb.save( filename = 'SACK_Droptail.xlsx' )
		if var=="SACK" and k=="RED":
                        a=[]
                        b=[]
                        c=[]
                        d=[]
                        a,b,c,d=throughput(var,k)
                        wb = Workbook()
                        ws = wb.create_sheet()
			#pylab.figure(1)
                        pylab.plot(a,b,color='red')
                        pylab.plot(c,d,color='green')
			pylab.legend(['Reno_TCP(RED)','Reno_CBR(RED)','SACK_TCP(RED)','SACK_CBR(RED)'],2)
			pylab.show()
                        #print a, len(a)
                        #print b,len(b)
                        #print c,len(c)
                      
                        for i in range(len(a)):
                                ws['A'+str(i + 1)].value = a[i]
                                ws['B'+str(i + 1)].value = b[i]
			for m in range(len(c)):
                                ws['D'+str(m + 1)].value = c[m]
                                ws['E'+str(m + 1)].value = d[m]
                                wb.save( filename = 'SACK_RED.xlsx' )




